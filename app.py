"""
JD Agentic AI Call Evaluation System — v2.0
=============================================
Single-file Streamlit app.

 #1  Speaker mapping     — keyword scoring + first-speaker fallback
 #2  Category context    — rubric injected per business category
 #3  Structured Gemini   — response_mime_type: application/json
 #4  Schema validation   — required keys checked, defaults injected
 #5  Crosstalk           — replaced with honest speaker-count check
 #6  Manual review flag  — set programmatically with reason list
 #7  Hallucination guard — fuzzy 70% word-match on transcript_quote
 #8  Language detection  — AssemblyAI language_detection:True
 #9  First response latency — computed from utterance timestamps
#10  Repeat detection    — programmatic SequenceMatcher similarity
#11  Unused imports      — none imported
#12  Temp file cleanup   — os.remove in finally block (always runs)
#13  Rate limiting       — exponential backoff + configurable sleep slider
#14  Batch processing    — sequential with live progress (fine for POC)
#15  Cost tracking       — token estimate + USD cost per call + total
#16  Category match      — per-category expected field map
#17  Off-script detect   — reference question list ratio
#18  Bot quality rubric  — weighted 4-dimension scoring (not subjective)
#19  Confidence scores   — 0.0-1.0 per field, threshold flagging
#20  Call metadata       — duration, latency, language, repeats in output
#21  No hardcoded keys   — os.getenv() only, never in source code
#22  No pyannote/HF      — fully replaced by AssemblyAI
#23  Logging             — Python logging module throughout
#24  Raw JSON sheet      — 6-sheet Excel report
#25  Accuracy sheet      — ground truth comparison template in Excel

Stack: AssemblyAI + Gemini 1.5 Flash + Streamlit + openpyxl
Deploy: Hugging Face Spaces (Streamlit SDK)
Secrets: ASSEMBLYAI_API_KEY, GROQ_API_KEY
"""

# ══════════════════════════════════════════════════════════════
# IMPORTS
# ══════════════════════════════════════════════════════════════

import os
import re
import json
import time
import logging
import tempfile
import difflib
from io import BytesIO
from datetime import datetime

import requests
import pandas as pd
import streamlit as st
from groq import Groq
TRANSLITERATION_AVAILABLE = True  # using built-in mapping, no external lib needed
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════
# LOGGING  — fix #23
# ══════════════════════════════════════════════════════════════

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S"
)
log = logging.getLogger("JD_Evaluator")


# ══════════════════════════════════════════════════════════════
# CONSTANTS
# ══════════════════════════════════════════════════════════════

ASSEMBLYAI_BASE_URL = "https://api.assemblyai.com/v2"

REQUIRED_FIELDS = [
    "Company Name", "Address", "Mobile Number",
    "Working Hours", "Contact Person", "Business Category"
]

# Reference questions the JD Bot SHOULD ask  — fix #17
# Includes actual Hinglish variants used in real JD calls
REFERENCE_QUESTIONS = [
    # Verification intent
    "verification", "verify", "verif", "confirm", "jankari",
    # Business name
    "company name", "aapka naam", "business ka naam", "nam shi hai",
    "kya aap", "nam btao", "business onr", "owner",
    # Address
    "address", "aapka pata", "location", "pata",
    # Mobile
    "mobile number", "phone number", "contact number", "mobile",
    # Working hours
    "working hours", "timing", "khula rehta", "band hota",
    "srvi", "service", "chalu", "bnd",
    # Contact person
    "contact person", "kaun hain aap", "owner ka naam", "onr",
    # Business category
    "business category", "kya karte hain", "kaun sa business",
    "restaurant", "hotel", "doctor", "clinic", "dukaan",
    # Common bot phrases in calls
    "shi hai", "sahi hai", "kya aapke", "aapke business",
    "basic details", "besik", "details", "ditels",
    "1 mint", "ek minute", "thoda time",
    "srvis", "avelebl", "opshn", "kliyr",
    "information", "inphormeshn", "nnbr", "gugl",
]

# Per-category expected fields  — fix #16
CATEGORY_FIELD_MAP = {
    "restaurant":   ["Company Name", "Address", "Mobile Number", "Working Hours", "Business Category"],
    "doctor":       ["Company Name", "Address", "Mobile Number", "Working Hours", "Contact Person"],
    "beauty":       ["Company Name", "Address", "Mobile Number", "Working Hours", "Contact Person"],
    "manufacturer": ["Company Name", "Address", "Mobile Number", "Contact Person", "Business Category"],
    "plumber":      ["Company Name", "Mobile Number", "Working Hours", "Contact Person"],
    "default":      REQUIRED_FIELDS
}

# Bot keywords for speaker identification  — fix #1
BOT_KEYWORDS = [
    # English
    "justdial", "verification", "verify", "confirm", "business details",
    "working hours", "listing", "contact person", "company name",
    "business category", "mobile number",
    # Hinglish
    "justdial se", "aapka naam", "aapka address", "aapka mobile",
    "timing", "khula rehta", "band hota", "kaun hain aap",
    "vyapar", "dukaan"
]

# Schema validation constants  — fix #4
REQUIRED_EVAL_KEYS = {
    "fields_confirmed": list,
    "bot_quality":      dict,
    "user_quality":     dict,
    "user_sentiment":   str,
    "call_outcome":     str,
    "off_script_flag":  bool,
    "manual_review":    bool,
    "overall_summary":  str,
}
REQUIRED_BOT_KEYS  = {"score", "reason", "followed_script", "asked_relevant_questions"}
REQUIRED_USER_KEYS = {"score", "reason", "cooperative", "clear_responses"}
VALID_SCORES       = {"Good", "Average", "Poor"}
VALID_SENTIMENTS   = {"Positive", "Neutral", "Negative"}
VALID_OUTCOMES     = {"Successful", "Partial", "Failed", "Incomplete"}


# ══════════════════════════════════════════════════════════════
# ① ASSEMBLYAI UTILITIES
# ══════════════════════════════════════════════════════════════

def upload_audio(file_path: str, api_key: str) -> str:
    log.info(f"Uploading: {file_path}")
    headers = {"authorization": api_key}
    with open(file_path, "rb") as f:
        r = requests.post(f"{ASSEMBLYAI_BASE_URL}/upload", headers=headers, data=f)
    r.raise_for_status()
    return r.json()["upload_url"]


def submit_transcription_job(audio_url: str, api_key: str) -> str:
    log.info("Submitting transcription job")
    headers = {"authorization": api_key, "content-type": "application/json"}
    payload = {
        "audio_url":          audio_url,
        "speaker_labels":     True,
        "language_detection": True,   # fix #8 — no forced Hindi
    }
    r = requests.post(f"{ASSEMBLYAI_BASE_URL}/transcript", json=payload, headers=headers)
    r.raise_for_status()
    return r.json()["id"]


def wait_for_completion(transcript_id: str, api_key: str, poll_interval: int = 3) -> dict:
    headers = {"authorization": api_key}
    while True:
        r = requests.get(f"{ASSEMBLYAI_BASE_URL}/transcript/{transcript_id}", headers=headers)
        r.raise_for_status()
        data   = r.json()
        status = data.get("status")
        log.info(f"AssemblyAI status: {status}")
        if status == "completed":
            return data
        if status == "error":
            raise Exception(f"AssemblyAI error: {data.get('error')}")
        time.sleep(poll_interval)


def transcribe_audio_assemblyai(file_path: str, api_key: str) -> dict:
    upload_url    = upload_audio(file_path, api_key)
    transcript_id = submit_transcription_job(upload_url, api_key)
    result        = wait_for_completion(transcript_id, api_key)
    # fix #8: audio_duration from AssemblyAI is already in seconds
    duration = round(result.get("audio_duration", 0), 2)
    log.info(f"Done. Duration={duration}s, Lang={result.get('language_code')}")
    return {
        "transcript":     result.get("text", ""),
        "utterances":     result.get("utterances", []),
        "language":       result.get("language_code", "unknown"),
        "audio_duration": duration,
    }


# ══════════════════════════════════════════════════════════════
# ② SPEAKER IDENTIFICATION  — fix #1
# ══════════════════════════════════════════════════════════════

def identify_bot_vendor(utterances: list) -> dict:
    """
    3-layer speaker identification — handles cases where vendor speaks first.

    Layer 1 — Keyword scoring (primary):
        Count Bot keywords in each speaker's turns.
        Highest score = Bot. This works regardless of who spoke first.

    Layer 2 — First utterance length check (tiebreak):
        If scores are tied, check first utterance word count.
        Bot opening is always long (10+ words: intro + purpose).
        Vendor pickup is always short (1-3 words: "haan", "ji boliye").
        Short first utterance → that speaker is Vendor, other is Bot.

    Layer 3 — Vendor keyword check (safety net):
        If first speaker's text contains vendor-only phrases
        (frustration, refusal, already-verified), they are the Vendor.
        Flip accordingly.
    """
    # Vendor-specific phrases — Bot would never say these
    VENDOR_KEYWORDS = [
        "already verified", "already verifid", "pehle ho gaya",
        "kyun call aa rha", "kyon call", "mat karo", "busy hoon",
        "nahi chahiye", "band karo", "kaun ho aap", "kaun bol rha",
        "aap kaun", "haan boliye", "ji boliye", "haan ji",
        "go go", "already done", "mera ho gaya", "ho gya",
    ]

    if not utterances:
        log.warning("No utterances found")
        return {}

    # ── Layer 1: Keyword scoring across ALL turns
    speaker_scores = {}
    speaker_word_counts = {}   # total words spoken per speaker

    for u in utterances:
        spk  = u.get("speaker")
        text = u.get("text", "").lower()
        words = len(text.split())
        speaker_scores.setdefault(spk, 0)
        speaker_word_counts[spk] = speaker_word_counts.get(spk, 0) + words
        for kw in BOT_KEYWORDS:
            if kw in text:
                speaker_scores[spk] += 1

    max_score  = max(speaker_scores.values())
    candidates = [s for s, sc in speaker_scores.items() if sc == max_score]

    # Clear winner from keyword scoring
    if len(candidates) == 1:
        bot_speaker = candidates[0]
        log.info(f"Layer 1 identified Bot: {bot_speaker} (score={max_score})")

    else:
        # ── Layer 2: Tiebreak by first utterance length
        first_spk  = utterances[0].get("speaker")
        first_text = utterances[0].get("text", "")
        first_word_count = len(first_text.split())

        if first_word_count <= 4:
            # Very short first utterance → vendor picked up ("haan", "ji boliye")
            # Bot is the OTHER speaker
            other_candidates = [s for s in candidates if s != first_spk]
            bot_speaker = other_candidates[0] if other_candidates else candidates[0]
            log.info(f"Layer 2 identified Bot: {bot_speaker} "
                     f"(first utterance only {first_word_count} words — likely vendor pickup)")
        else:
            # Long first utterance → Bot introduced itself first
            bot_speaker = first_spk if first_spk in candidates else candidates[0]
            log.info(f"Layer 2 identified Bot: {bot_speaker} "
                     f"(first utterance {first_word_count} words — likely Bot intro)")

    # ── Layer 3: Safety net — check if Bot candidate says vendor phrases
    bot_all_text = " ".join(
        u.get("text", "").lower()
        for u in utterances
        if u.get("speaker") == bot_speaker
    )
    vendor_hits = sum(1 for kw in VENDOR_KEYWORDS if kw in bot_all_text)

    if vendor_hits >= 2:
        # Bot candidate is saying too many vendor phrases — flip the mapping
        all_speakers = list(speaker_scores.keys())
        other_speaker = next((s for s in all_speakers if s != bot_speaker), None)
        if other_speaker:
            log.warning(f"Layer 3 FLIP: {bot_speaker} had {vendor_hits} vendor phrases "
                        f"— reassigning Bot to {other_speaker}")
            bot_speaker = other_speaker

    mapping = {s: ("Bot" if s == bot_speaker else "Vendor") for s in speaker_scores}
    log.info(f"Final speaker map: {mapping}")
    return mapping


def to_hinglish(text: str) -> str:
    """
    Convert Devanagari Hindi to natural Roman script (Hinglish).
    e.g. हेलो → hello, बिज़नेस → biznes, फ़ोन → fon
    """
    if not text:
        return text

    result = text

    # Step 0: Decomposed nukta FIRST (U+093C after base consonant)
    NUKTA = '\u093C'
    decomposed_nukta = {
        'ज'+NUKTA:'z', 'फ'+NUKTA:'f', 'ख'+NUKTA:'kh',
        'ग'+NUKTA:'gh', 'क'+NUKTA:'q', 'ड'+NUKTA:'r', 'ढ'+NUKTA:'rh',
    }
    for combo, rom in decomposed_nukta.items():
        result = result.replace(combo, rom)

    # Step 1: Full Devanagari character mapping
    mapping = {
        'क्ष':'ksh','त्र':'tr','ज्ञ':'gya',
        'अ':'a','आ':'aa','इ':'i','ई':'ee','उ':'u','ऊ':'oo',
        'ए':'e','ऐ':'ai','ओ':'o','औ':'au','ऋ':'ri','ऑ':'o',
        'क':'k','ख':'kh','ग':'g','घ':'gh','ङ':'n',
        'च':'ch','छ':'chh','ज':'j','झ':'jh','ञ':'n',
        'ट':'t','ठ':'th','ड':'d','ढ':'dh','ण':'n',
        'त':'t','थ':'th','द':'d','ध':'dh','न':'n',
        'प':'p','फ':'ph','ब':'b','भ':'bh','म':'m',
        'य':'y','र':'r','ल':'l','व':'v',
        'श':'sh','ष':'sh','स':'s','ह':'h',
        'ड़':'r','ढ़':'rh','ज़':'z','फ़':'f','ख़':'kh','ग़':'gh','क़':'q',
        'ा':'a','ि':'i','ी':'i','ु':'u','ू':'u',
        'े':'e','ै':'ai','ो':'o','ौ':'au','ॉ':'o',
        'ं':'n','ँ':'n','ः':'h','्':'',
        '।':'. ','₹':'Rs',
        '\u200c':'','\u200d':'','\u200b':'',
    }
    for dev, rom in mapping.items():
        result = result.replace(dev, rom)

    # Step 2: Fix common phonetic spellings of English loanwords
    result = result.lower()
    word_fixes = {
        'veriphikeshn':'verification','veriphikeshan':'verification',
        'vebsait':'website','knphrm':'confirm','knphirm':'confirm',
        'ditels':'details','bisnes':'business','bijnes':'business',
        'knpni':'company','kmpni':'company','aadres':'address',
        'mobaail':'mobile','mobal':'mobile',
        'helo':'hello','nmste':'namaste',
        'jlday':'jaldi','jldi':'jaldi',
        'sirph':'sirf','lgega':'lagega',
        'smjh':'samajh','puri trh':'puri tarah',
    }
    for wrong, right in word_fixes.items():
        result = result.replace(wrong, right)

    return result
def format_transcript(utterances: list, speaker_map: dict) -> str:
    lines = []
    for u in utterances:
        label = speaker_map.get(u.get("speaker"), u.get("speaker"))
        text  = to_hinglish(u.get("text", ""))   # Devanagari → Hinglish Roman
        lines.append(f"{label}: {text}")
    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════
# ③ PROGRAMMATIC METRICS  — fix #9, #10, #15, #17
# ══════════════════════════════════════════════════════════════

def compute_first_response_latency(utterances: list, speaker_map: dict) -> float:
    """
    fix #9: Time from end of Bot's first utterance to start of Vendor's first utterance.
    AssemblyAI gives start/end in MILLISECONDS — convert to seconds.
    Returns -1.0 if not computable.
    """
    bot_end      = None
    vendor_start = None
    for u in utterances:
        role    = speaker_map.get(u.get("speaker"), "")
        start_s = u.get("start", 0) / 1000.0
        end_s   = u.get("end",   0) / 1000.0
        if role == "Bot" and bot_end is None:
            bot_end = end_s
        if role == "Vendor" and vendor_start is None:
            vendor_start = start_s
        if bot_end is not None and vendor_start is not None:
            break

    if bot_end is not None and vendor_start is not None:
        latency = round(max(vendor_start - bot_end, 0.0), 2)
        log.info(f"First response latency: {latency}s")
        return latency
    log.warning("Could not compute latency")
    return -1.0


def compute_repeat_question_count(utterances: list, speaker_map: dict, threshold: float = 0.75) -> int:
    """
    fix #10: Programmatic repeat detection using SequenceMatcher.
    Compares all Bot utterances pairwise.
    If similarity >= threshold, counts as a repeat.
    """
    bot_texts = [
        to_hinglish(u.get("text", "")).lower().strip()   # ← transliterate FIRST
        for u in utterances
        if speaker_map.get(u.get("speaker")) == "Bot" and len(u.get("text", "")) > 10
    ]
    repeat_count = 0
    seen = []
    for text in bot_texts:
        for prev in seen:
            ratio = difflib.SequenceMatcher(None, text, prev).ratio()
            if ratio >= threshold:
                repeat_count += 1
                log.info(f"Repeat found (sim={ratio:.2f}): '{text[:60]}'")
                break
        seen.append(text)
    return repeat_count


def compute_off_script_ratio(utterances: list, speaker_map: dict) -> float:
    """
    fix #17: Fraction of Bot turns that contain at least one reference question keyword.
    Applies to_hinglish() FIRST so Hindi text matches Roman keywords correctly.
    Low ratio = bot went off-script.
    """
    bot_texts = [
        to_hinglish(u.get("text", "")).lower()   # ← transliterate FIRST
        for u in utterances
        if speaker_map.get(u.get("speaker")) == "Bot"
    ]
    if not bot_texts:
        return 0.0
    on_script = sum(1 for t in bot_texts if any(kw in t for kw in REFERENCE_QUESTIONS))
    ratio = round(on_script / len(bot_texts), 2)
    log.info(f"On-script ratio: {ratio} ({on_script}/{len(bot_texts)} bot turns matched)")
    return ratio


def estimate_token_cost(transcript: str) -> dict:
    """
    fix #15: Rough cost estimate.
    Gemini 1.5 Flash: ~$0.075 per 1M input tokens. 1 token ~ 4 chars.
    """
    approx_tokens = len(transcript) // 4
    cost_usd      = round(approx_tokens * 0.59 / 1_000_000, 6)   # Groq LLaMA 3.3 70B = $0.59 per 1M tokens
    return {"approx_tokens": approx_tokens, "estimated_cost_usd": cost_usd}


# ══════════════════════════════════════════════════════════════
# ④ MANUAL REVIEW DETECTION  — fix #6
# ══════════════════════════════════════════════════════════════

def detect_manual_review_conditions(
    assemblyai_result: dict,
    repeat_count: int,
    off_script_ratio: float
) -> dict:
    reasons    = []
    duration   = assemblyai_result.get("audio_duration", 0)
    utterances = assemblyai_result.get("utterances", [])
    speakers   = {u.get("speaker") for u in utterances}

    if duration < 30:
        reasons.append("Call shorter than 30 seconds")
    if len(speakers) > 2:
        reasons.append("More than 2 speakers detected (possible crosstalk)")  # fix #5
    if not utterances:
        reasons.append("No transcript generated")
    if repeat_count >= 2:
        reasons.append(f"Bot repeated questions {repeat_count} time(s)")
    if 0.0 <= off_script_ratio < 0.1:
        # Only flag if truly no script keywords found AND call is long enough
        # Short ratio is often a keyword-matching limitation, not actual off-script
        if duration > 30:
            reasons.append(f"Possible off-script (keyword ratio: {off_script_ratio}) — verify manually")

    return {"manual_review": len(reasons) > 0, "review_reasons": reasons}


# ══════════════════════════════════════════════════════════════
# ⑤ LLM PROMPT  — fix #2, #16, #17, #18
# ══════════════════════════════════════════════════════════════

def get_category_rubric(business_category: str) -> str:
    """fix #16: Return expected fields and rubric for this category."""
    cat = (business_category or "").lower()
    for key, fields in CATEGORY_FIELD_MAP.items():
        if key in cat:
            return f"Expected fields for '{key}': {', '.join(fields)}"
    return f"Expected fields (default): {', '.join(CATEGORY_FIELD_MAP['default'])}"


def build_evaluation_prompt(
    transcript: str,
    business_category: str,
    repeat_count: int,
    off_script_ratio: float,
    first_latency: float
) -> str:
    """
    fix #2 #16 #17 #18: Category context + reference questions +
    weighted scoring rubric + pre-computed metrics injected.
    """
    cat_rubric = get_category_rubric(business_category)
    bc         = business_category or "Unknown"

    return f"""
You are an expert Call Quality Analyst for JustDial.
Evaluate this phone conversation between a JD Agentic AI Bot and a business vendor.
Transcript language: Hinglish (Hindi in Roman script + English).
Bot goal: collect and verify vendor business details.

BUSINESS CATEGORY: {bc}
{cat_rubric}

TRANSCRIPT:
{transcript}

PRE-COMPUTED METRICS (use these exactly — do not re-estimate):
- first_response_latency_seconds: {first_latency}
- repeat_question_count: {repeat_count}
- on_script_ratio (0.0=off-script, 1.0=fully on-script): {off_script_ratio}

CRITICAL RULE FOR BOT QUALITY EVALUATION:
If the vendor refused to cooperate, got angry, or disconnected early —
that is a USER/VENDOR failure, NOT a bot failure.
Judge bot quality ONLY on what the bot did:
- Did it introduce the call purpose? (verification call)
- Did it ask about business name, address, timings, contact?
- Did it stay polite and professional even when vendor was rude?
- Did it try to re-engage after objections?
If yes to most — Bot Quality = Good, regardless of call outcome.
NEVER penalize the bot for vendor behavior.

TASK 1 — FIELD VERIFICATION
For each field — Company Name, Address, Mobile Number, Working Hours,
Contact Person, Business Category — provide:
  status:           Confirmed | Partial | Not Confirmed
  extracted_value:  exact value from transcript, null if not found
  transcript_quote: verbatim line from the transcript that supports it
  confidence:       0.0 to 1.0
Rules: Do NOT hallucinate. Only Confirmed if clearly stated.
Every Confirmed field MUST have a non-empty transcript_quote.

TASK 2 — BOT QUALITY  (fix #18 weighted rubric)
Score each dimension 1-5:
  script_adherence_score (30%):   Did bot ask all expected fields for this category?
  question_relevance_score (30%): Were questions appropriate for vendor type?
  no_repetition_score (20%):      Avoid repeating? Use repeat_count={repeat_count} provided above.
  professionalism_score (20%):    Professional, patient, clear tone?
Weighted average: Good if >= 3.5 | Average if 2.5-3.4 | Poor if < 2.5
Return score + reason (2 sentences max).

TASK 3 — USER QUALITY
Evaluate: cooperation, clarity, frustration, early disconnect, relevance.
Score: Good | Average | Poor. Reason: 2 sentences max.

TASK 4 — SENTIMENT
Vendor overall sentiment: Positive | Neutral | Negative

TASK 5 — CALL OUTCOME
Successful | Partial | Failed | Incomplete

TASK 6 — OFF-SCRIPT FLAG
Set off_script_flag=true if bot asked questions unrelated to business verification.
Reference: on_script_ratio={off_script_ratio} (< 0.5 strongly suggests off-script).

TASK 7 — SUMMARY
2-3 sentence plain-English summary for a manager.

STRICT RULES:
1. Return ONLY valid JSON — no markdown, no explanation.
2. Never invent values not present in the transcript.
3. Use the pre-computed metrics above — do not override them.

OUTPUT JSON:
{{
  "fields_confirmed": [
    {{
      "field": "",
      "status": "",
      "extracted_value": "",
      "transcript_quote": "",
      "confidence": 0.0
    }}
  ],
  "bot_quality": {{
    "score": "",
    "script_adherence_score": 0,
    "question_relevance_score": 0,
    "no_repetition_score": 0,
    "professionalism_score": 0,
    "followed_script": true,
    "asked_relevant_questions": true,
    "unnecessary_questions": false,
    "handled_objections": true,
    "reason": ""
  }},
  "user_quality": {{
    "score": "",
    "cooperative": true,
    "clear_responses": true,
    "showed_frustration": false,
    "disconnected_early": false,
    "reason": ""
  }},
  "user_sentiment": "",
  "call_outcome": "",
  "off_script_flag": false,
  "manual_review": false,
  "overall_summary": ""
}}
"""


# ══════════════════════════════════════════════════════════════
# ⑥ SCHEMA VALIDATION  — fix #4
# ══════════════════════════════════════════════════════════════

def numeric_score_to_label(score) -> str:
    """
    Convert Groq's numeric weighted score to Good/Average/Poor label.
    Groq sometimes returns 3.5 instead of 'Good' — this handles it cleanly.
    """
    try:
        val = float(score)
        if val >= 3.5:
            return "Good"
        elif val >= 2.5:
            return "Average"
        else:
            return "Poor"
    except (TypeError, ValueError):
        return None   # not a number — let caller handle


def validate_and_repair_schema(ev: dict) -> tuple:
    """
    fix #4: Validate Groq JSON against required schema.
    Injects safe defaults for missing keys — never crashes.
    Converts numeric scores (3.5) to labels (Good) automatically.
    Returns (repaired_dict, warnings_list).
    """
    warnings_list = []

    for key, expected_type in REQUIRED_EVAL_KEYS.items():
        if key not in ev:
            warnings_list.append(f"Missing key '{key}' — injected default")
            if expected_type == list:   ev[key] = []
            elif expected_type == dict: ev[key] = {}
            elif expected_type == str:  ev[key] = "Unknown"
            elif expected_type == bool: ev[key] = False

    bq = ev.get("bot_quality", {})
    for k in REQUIRED_BOT_KEYS:
        if k not in bq:
            warnings_list.append(f"bot_quality missing '{k}'")
            bq[k] = "Unknown" if k in ("score", "reason") else False
    if bq.get("score") not in VALID_SCORES:
        # Try converting numeric score first
        converted = numeric_score_to_label(bq.get("score"))
        if converted:
            log.info(f"Bot score {bq.get('score')} → {converted}")
            bq["score"] = converted
        else:
            warnings_list.append(f"Invalid bot score '{bq.get('score')}' — set to Average")
            bq["score"] = "Average"

    uq = ev.get("user_quality", {})
    for k in REQUIRED_USER_KEYS:
        if k not in uq:
            warnings_list.append(f"user_quality missing '{k}'")
            uq[k] = "Unknown" if k in ("score", "reason") else False
    if uq.get("score") not in VALID_SCORES:
        # Try converting numeric score first
        converted = numeric_score_to_label(uq.get("score"))
        if converted:
            log.info(f"User score {uq.get('score')} → {converted}")
            uq["score"] = converted
        else:
            warnings_list.append(f"Invalid user score '{uq.get('score')}' — set to Average")
            uq["score"] = "Average"

    if ev.get("user_sentiment") not in VALID_SENTIMENTS:
        warnings_list.append(f"Invalid sentiment — set to Neutral")
        ev["user_sentiment"] = "Neutral"
    if ev.get("call_outcome") not in VALID_OUTCOMES:
        warnings_list.append(f"Invalid outcome — set to Incomplete")
        ev["call_outcome"] = "Incomplete"

    for f in ev.get("fields_confirmed", []):
        try:
            f["confidence"] = float(f.get("confidence", 0))
        except (TypeError, ValueError):
            f["confidence"] = 0.0
            warnings_list.append(f"Bad confidence for '{f.get('field')}' — set to 0.0")

    if warnings_list:
        log.warning(f"Schema repairs ({len(warnings_list)}): {warnings_list}")

    return ev, warnings_list


# ══════════════════════════════════════════════════════════════
# ⑦ HALLUCINATION VALIDATOR  — fix #7
# ══════════════════════════════════════════════════════════════

def normalize_text(text) -> str:
    if not text:
        return ""
    return re.sub(r"\s+", " ", str(text).lower().strip())


def quote_exists_in_transcript(quote: str, transcript: str) -> bool:
    """Fuzzy: >= 70% of quote words must appear in transcript."""
    words = normalize_text(quote).split()
    trans = normalize_text(transcript)
    if not words:
        return False
    hits = sum(1 for w in words if w in trans)
    return (hits / len(words)) >= 0.70


def validate_fields_against_transcript(evaluation: dict, transcript: str) -> dict:
    """fix #7: Verify each confirmed field's quote against actual transcript."""
    all_reasons = []
    for field in evaluation.get("fields_confirmed", []):
        quote      = field.get("transcript_quote", "")
        confidence = field.get("confidence", 0)
        reasons    = []

        if field.get("status") == "Confirmed":
            if not quote_exists_in_transcript(quote, transcript):
                reasons.append(f"{field.get('field')}: quote not found (possible hallucination)")
                field["confidence"] = min(float(confidence), 0.3)

        if float(field.get("confidence", 0)) < 0.50:
            reasons.append(f"{field.get('field')}: low confidence ({field.get('confidence')})")

        field["validation_warnings"] = reasons
        all_reasons.extend(reasons)

    found   = [f.get("field") for f in evaluation.get("fields_confirmed", [])]
    missing = [f for f in REQUIRED_FIELDS if f not in found]
    if missing:
        all_reasons.append(f"Missing fields: {', '.join(missing)}")

    if all_reasons:
        evaluation["manual_review"] = True
        evaluation.setdefault("manual_review_reason", [])
        evaluation["manual_review_reason"].extend(list(set(all_reasons)))

    return evaluation


# ══════════════════════════════════════════════════════════════
# ⑧ GROQ CALL WITH RETRY  — replaces Gemini (fix #3, #13)
# ══════════════════════════════════════════════════════════════

def call_gemini_with_retry(groq_client, prompt: str, max_retries: int = 3) -> dict:
    """
    Calls Groq API (llama-3.3-70b) instead of Gemini.
    Groq free tier: 14,400 requests/day — no quota issues.
    """
    for attempt in range(1, max_retries + 1):
        try:
            log.info(f"Groq attempt {attempt}")
            response = groq_client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[
                    {
                        "role": "system",
                        "content": "You are an expert Call Quality Analyst. Always respond with valid JSON only. No markdown, no explanation, just raw JSON."
                    },
                    {
                        "role": "user",
                        "content": prompt
                    }
                ],
                temperature=0.2,
                response_format={"type": "json_object"},
                max_tokens=4000
            )
            raw = response.choices[0].message.content.strip()
            if raw.startswith("```"):
                raw = raw.split("```")[1]
                if raw.startswith("json"):
                    raw = raw[4:]
            result = json.loads(raw)
            log.info("Groq response parsed OK")
            return result
        except json.JSONDecodeError as e:
            log.error(f"JSON parse error attempt {attempt}: {e}")
        except Exception as e:
            log.error(f"Groq error attempt {attempt}: {e}")
        wait = 2 ** attempt
        log.info(f"Retry in {wait}s")
        time.sleep(wait)

    raise Exception("Groq evaluation failed after all retries")


# ══════════════════════════════════════════════════════════════
# ⑨ EXCEL EXPORT  — fix #24, #25
# ══════════════════════════════════════════════════════════════

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
HEADER_FONT  = Font(color="FFFFFF", bold=True)
GOOD_FILL    = PatternFill("solid", fgColor="C6EFCE")
AVG_FILL     = PatternFill("solid", fgColor="FFEB9C")
POOR_FILL    = PatternFill("solid", fgColor="FFC7CE")
NEUTRAL_FILL = PatternFill("solid", fgColor="DDEBF7")
FLAG_FILL    = PatternFill("solid", fgColor="FCE4D6")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)
COLOR_MAP = {
    "Good": GOOD_FILL, "Positive": GOOD_FILL, "Successful": GOOD_FILL, "Confirmed": GOOD_FILL,
    "Excellent": GOOD_FILL,
    "Average": AVG_FILL, "Neutral": NEUTRAL_FILL, "Partial": AVG_FILL, "Incomplete": AVG_FILL,
    "Poor": POOR_FILL, "Negative": POOR_FILL, "Failed": POOR_FILL, "Not Confirmed": POOR_FILL,
    "Yes": FLAG_FILL, "Yes ": FLAG_FILL,
}


def style_sheet(ws):
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border    = THIN_BORDER
            val = str(cell.value) if cell.value is not None else ""
            if val in COLOR_MAP:
                cell.fill = COLOR_MAP[val]
    for col in ws.columns:
        mx = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 4, 60)


def create_excel_report(results: list) -> bytes:
    """
    6-sheet Excel:
    1. Summary        — one row per call
    2. Fields Detail  — one row per field per call
    3. Metrics        — latency, repeats, cost, language (#9/#10/#15/#20)
    4. Transcripts    — full transcript per call
    5. Raw JSON       — full Gemini output per call (#24)
    6. Accuracy       — ground truth comparison template (#25)
    """
    summary_rows, field_rows, metric_rows = [], [], []
    transcript_rows, raw_json_rows, accuracy_rows = [], [], []

    for r in results:
        fn  = r.get("file_name", "")
        tr  = r.get("transcript", "")
        ev  = r.get("evaluation", {})
        met = r.get("metrics", {})

        confirmed_count = sum(
            1 for f in ev.get("fields_confirmed", [])
            if f.get("status") == "Confirmed"
        )

        partial_count = sum(
            1 for f in ev.get("fields_confirmed", [])
            if f.get("status") == "Partial"
        )

        total_fields = len(REQUIRED_FIELDS)

        # Data Collection Score — judges end result regardless of why
        if confirmed_count >= 5:
            data_collection_score = "Excellent"
            data_collection_note  = f"{confirmed_count}/{total_fields} fields confirmed"
        elif confirmed_count >= 3:
            data_collection_score = "Good"
            data_collection_note  = f"{confirmed_count}/{total_fields} fields confirmed"
        elif confirmed_count >= 1 or partial_count >= 2:
            data_collection_score = "Partial"
            data_collection_note  = f"{confirmed_count} confirmed, {partial_count} partial"
        else:
            data_collection_score = "Failed"
            data_collection_note  = "No fields collected"

        summary_rows.append({
            "File Name":              fn,
            "Call Outcome":           ev.get("call_outcome"),
            "Bot Quality":            ev.get("bot_quality", {}).get("score"),
            "Bot Reason":             ev.get("bot_quality", {}).get("reason"),
            "User Quality":           ev.get("user_quality", {}).get("score"),
            "User Reason":            ev.get("user_quality", {}).get("reason"),
            "User Sentiment":         ev.get("user_sentiment"),
            "Fields Confirmed":       f"{confirmed_count}/{total_fields}",
            "Data Collection Score":  data_collection_score,
            "Data Collection Note":   data_collection_note,
            "Off Script":             "Yes" if ev.get("off_script_flag") else "No",
            "Manual Review":          "Yes" if ev.get("manual_review") else "No",
            "Review Reasons":         "; ".join(ev.get("manual_review_reason", [])),
            "Schema Warnings":        "; ".join(r.get("schema_warnings", [])),
            "Overall Summary":        ev.get("overall_summary"),
            "Full Transcript":        tr,
        })

        for f in ev.get("fields_confirmed", []):
            field_rows.append({
                "File Name":          fn,
                "Field":              f.get("field"),
                "Status":             f.get("status"),
                "Extracted Value":    f.get("extracted_value"),
                "Confidence":         f.get("confidence"),
                "Transcript Quote":   f.get("transcript_quote"),
                "Validation Warnings":"; ".join(f.get("validation_warnings", [])),
            })

        metric_rows.append({
            "File Name":                   fn,
            "Duration (s)":                met.get("duration_s"),
            "Language":                    met.get("language"),
            "First Response Latency (s)":  met.get("first_latency_s"),
            "Repeat Question Count":       met.get("repeat_count"),
            "On-Script Ratio":             met.get("off_script_ratio"),
            "Approx Tokens":               met.get("approx_tokens"),
            "Estimated Cost (USD)":        met.get("estimated_cost_usd"),
            "Processed At":                met.get("processed_at"),
        })

        transcript_rows.append({"File Name": fn, "Transcript": tr})

        raw_json_rows.append({
            "File Name": fn,
            "Raw JSON":  json.dumps(ev, ensure_ascii=False, indent=2)
        })

        # Ground truth template — human fills in the blank columns  (#25)
        for f in ev.get("fields_confirmed", []):
            accuracy_rows.append({
                "File Name":          fn,
                "Field":              f.get("field"),
                "AI Status":          f.get("status"),
                "AI Extracted Value": f.get("extracted_value"),
                "Ground Truth Value": "",    # human fills
                "Ground Truth Status":"",    # human fills
                "Match (Y/N)":        "",    # human fills
                "Notes":              "",
            })

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows).to_excel(    writer, sheet_name="Summary",       index=False)
        pd.DataFrame(field_rows).to_excel(      writer, sheet_name="Fields Detail", index=False)
        pd.DataFrame(metric_rows).to_excel(     writer, sheet_name="Metrics",       index=False)
        pd.DataFrame(transcript_rows).to_excel( writer, sheet_name="Transcripts",   index=False)
        pd.DataFrame(raw_json_rows).to_excel(   writer, sheet_name="Raw JSON",      index=False)
        pd.DataFrame(accuracy_rows).to_excel(   writer, sheet_name="Accuracy",      index=False)

    output.seek(0)
    wb = load_workbook(output)
    for sn in wb.sheetnames:
        style_sheet(wb[sn])

    final = BytesIO()
    wb.save(final)
    final.seek(0)
    return final.getvalue()


# ══════════════════════════════════════════════════════════════
# ⑩ STREAMLIT UI
# ══════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="JD Agentic AI Call Evaluator",
    page_icon="📞",
    layout="wide"
)

# fix #21 — keys from env only, never hardcoded
ASSEMBLYAI_API_KEY = os.getenv("ASSEMBLYAI_API_KEY", "")
GROQ_API_KEY       = os.getenv("GROQ_API_KEY", "")

if not ASSEMBLYAI_API_KEY:
    st.error("ASSEMBLYAI_API_KEY not set. Add it in HF Space → Settings → Repository Secrets.")
    st.stop()
if not GROQ_API_KEY:
    st.error("GROQ_API_KEY not set. Add it in HF Space → Settings → Repository Secrets.")
    st.stop()

gemini_client = Groq(api_key=GROQ_API_KEY)  # using Groq instead of Gemini

# ── Header
st.title("📞 JD Agentic AI Call Evaluation System")
st.caption("v2.0 — All 25 issues addressed")
st.divider()

# ── Sidebar
with st.sidebar:
    st.header("Configuration")

    business_category = st.text_input(
        "Business Category",
        placeholder="e.g. Restaurant / Doctor / Beauty"
    )

    gemini_delay = st.slider(  # fix #13
        "Delay between API calls (s)", 2, 15, 4,
        help="Increase if hitting free-tier rate limits"
    )

    st.markdown("---")
    st.markdown("**Expected fields:**")
    for f in REQUIRED_FIELDS:
        st.markdown(f"• {f}")

    st.markdown("---")
    st.markdown("**Reference script keywords:**")
    for q in REFERENCE_QUESTIONS[:8]:
        st.markdown(f"• {q}")

# ── File upload
uploaded_files = st.file_uploader(
    "Upload Call Recordings",
    type=["mp3", "wav", "m4a", "ogg"],
    accept_multiple_files=True
)

if uploaded_files:
    est_time = len(uploaded_files) * 45
    st.info(f"{len(uploaded_files)} file(s) ready — estimated time: ~{est_time}s")

# ── Process button
if st.button("Evaluate Calls", type="primary", disabled=not uploaded_files):

    results  = []
    progress = st.progress(0, text="Starting…")

    for idx, uf in enumerate(uploaded_files):
        progress.progress(idx / len(uploaded_files), text=f"Processing {uf.name}…")

        with st.status(f"{uf.name}", expanded=True) as status:
            temp_path = None
            try:
                # Save temp file
                suffix = "." + uf.name.split(".")[-1]
                with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                    tmp.write(uf.read())
                    temp_path = tmp.name

                # Step 1: AssemblyAI transcription + diarization
                st.write("Transcribing with AssemblyAI…")
                transcript_data = transcribe_audio_assemblyai(temp_path, ASSEMBLYAI_API_KEY)

                # Step 2: Speaker identification  (fix #1)
                st.write("Identifying speakers…")
                speaker_map          = identify_bot_vendor(transcript_data["utterances"])
                formatted_transcript = format_transcript(transcript_data["utterances"], speaker_map)

                # Step 3: Programmatic metrics  (fix #9 #10 #15 #17)
                st.write("Computing call metrics…")
                first_latency    = compute_first_response_latency(transcript_data["utterances"], speaker_map)
                repeat_count     = compute_repeat_question_count(transcript_data["utterances"], speaker_map)
                off_script_ratio = compute_off_script_ratio(transcript_data["utterances"], speaker_map)
                cost_info        = estimate_token_cost(formatted_transcript)

                # Step 4: Gemini evaluation  (fix #2 #3 #13 #18)
                st.write("Evaluating with Groq LLaMA 3.3…")
                prompt     = build_evaluation_prompt(
                    formatted_transcript, business_category,
                    repeat_count, off_script_ratio, first_latency
                )
                evaluation = call_gemini_with_retry(gemini_client, prompt)

                # Step 5: Schema validation  (fix #4)
                st.write("Validating schema…")
                evaluation, schema_warnings = validate_and_repair_schema(evaluation)

                # Step 6: Hallucination check  (fix #7)
                st.write("Checking for hallucinations…")
                evaluation = validate_fields_against_transcript(evaluation, formatted_transcript)

                # Step 7: Manual review flags  (fix #6)
                review = detect_manual_review_conditions(transcript_data, repeat_count, off_script_ratio)
                if review["manual_review"]:
                    evaluation["manual_review"] = True
                    evaluation.setdefault("manual_review_reason", [])
                    evaluation["manual_review_reason"].extend(review["review_reasons"])

                # Inject programmatic metrics  (fix #20)
                evaluation["first_response_latency_s"] = first_latency
                evaluation["repeat_question_count"]    = repeat_count
                evaluation["on_script_ratio"]          = off_script_ratio

                results.append({
                    "file_name":       uf.name,
                    "transcript":      formatted_transcript,
                    "evaluation":      evaluation,
                    "schema_warnings": schema_warnings,
                    "metrics": {
                        "duration_s":         transcript_data["audio_duration"],
                        "language":           transcript_data["language"],
                        "first_latency_s":    first_latency,
                        "repeat_count":       repeat_count,
                        "off_script_ratio":   off_script_ratio,
                        "approx_tokens":      cost_info["approx_tokens"],
                        "estimated_cost_usd": cost_info["estimated_cost_usd"],
                        "processed_at":       datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    }
                })

                status.update(label=f"Done — {uf.name}", state="complete", expanded=False)
                log.info(f"Completed: {uf.name}")

            except Exception as e:
                log.error(f"Failed {uf.name}: {e}")
                status.update(label=f"Error — {uf.name}: {str(e)}", state="error", expanded=False)
                st.error(str(e))

            finally:
                # fix #12 — always clean up temp file
                if temp_path and os.path.exists(temp_path):
                    os.remove(temp_path)
                    log.info(f"Deleted temp: {temp_path}")

            # fix #13 — configurable rate limit delay
            if idx < len(uploaded_files) - 1:
                time.sleep(gemini_delay)

        progress.progress((idx + 1) / len(uploaded_files), text=f"{idx+1}/{len(uploaded_files)} done")

    progress.progress(1.0, text="All done!")

    # ── Results
    if results:
        st.success(f"{len(results)}/{len(uploaded_files)} calls evaluated successfully.")

        # KPI row  (fix #15 cost tracking)
        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("Total Calls",   len(results))
        c2.metric("Successful",    sum(1 for r in results if r["evaluation"].get("call_outcome") == "Successful"))
        c3.metric("Manual Review", sum(1 for r in results if r["evaluation"].get("manual_review")))
        c4.metric("Off Script",    sum(1 for r in results if r["evaluation"].get("off_script_flag")))
        total_cost = sum(r["metrics"]["estimated_cost_usd"] for r in results)
        c5.metric("Est. Cost (USD)", f"${total_cost:.4f}")

        st.divider()

        tab1, tab2, tab3, tab4, tab5 = st.tabs(
            ["Summary", "Fields", "Metrics", "Transcript", "Raw JSON"]
        )

        with tab1:
            rows = []
            for r in results:
                ev = r["evaluation"]
                confirmed = sum(1 for f in ev.get("fields_confirmed", []) if f.get("status") == "Confirmed")
                partial   = sum(1 for f in ev.get("fields_confirmed", []) if f.get("status") == "Partial")
                if confirmed >= 5:        dcs = "Excellent"
                elif confirmed >= 3:      dcs = "Good"
                elif confirmed >= 1 or partial >= 2: dcs = "Partial"
                else:                     dcs = "Failed"
                rows.append({
                    "File":                   r["file_name"],
                    "Outcome":                ev.get("call_outcome"),
                    "Bot Quality":            ev.get("bot_quality", {}).get("score"),
                    "User Quality":           ev.get("user_quality", {}).get("score"),
                    "Sentiment":              ev.get("user_sentiment"),
                    "Fields":                 f"{confirmed}/{len(REQUIRED_FIELDS)}",
                    "Data Collection Score":  dcs,
                    "Repeats":                ev.get("repeat_question_count", 0),
                    "Off Script":             "Yes" if ev.get("off_script_flag") else "No",
                    "Manual Review":          "Yes" if ev.get("manual_review") else "No",
                    "Summary":                ev.get("overall_summary", ""),
                })
            st.dataframe(pd.DataFrame(rows), use_container_width=True)

        with tab2:
            for r in results:
                st.subheader(r["file_name"])
                fd = [
                    {
                        "Field":     f.get("field"),
                        "Status":    f.get("status"),
                        "Value":     f.get("extracted_value"),
                        "Confidence":f.get("confidence"),
                        "Quote":     f.get("transcript_quote"),
                        "Warnings":  "; ".join(f.get("validation_warnings", [])),
                    }
                    for f in r["evaluation"].get("fields_confirmed", [])
                ]
                st.dataframe(pd.DataFrame(fd), use_container_width=True)

        with tab3:
            for r in results:
                st.subheader(r["file_name"])
                m = r["metrics"]
                m1, m2, m3, m4, m5 = st.columns(5)
                m1.metric("Duration",       f"{m['duration_s']}s")
                m2.metric("Latency",        f"{m['first_latency_s']}s")
                m3.metric("Repeats",        m["repeat_count"])
                m4.metric("On-Script",      m["off_script_ratio"])
                m5.metric("Est. Cost",      f"${m['estimated_cost_usd']:.4f}")
                if r["schema_warnings"]:
                    st.warning("Schema warnings: " + " | ".join(r["schema_warnings"]))

        with tab4:
            for r in results:
                st.subheader(r["file_name"])
                st.text_area(
                    "Transcript", r["transcript"],
                    height=250, key=f"tr_{r['file_name']}"
                )

        with tab5:
            for r in results:
                st.subheader(r["file_name"])
                st.json(r["evaluation"])

        # Download
        st.divider()
        excel_bytes = create_excel_report(results)
        ts          = datetime.now().strftime("%Y%m%d_%H%M")
        st.download_button(
            label="Download Excel Report (6 sheets)",
            data=excel_bytes,
            file_name=f"JD_Call_Evaluation_{ts}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
